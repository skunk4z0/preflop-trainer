from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# --- debug: config import & REF_COLOR_CELLS ---
import sys
ROOT = Path(__file__).resolve().parents[1]  # C:\MyPokerApp
sys.path.insert(0, str(ROOT))

from config import REF_COLOR_CELLS  # ← これが使える前提にする（try/exceptは消す）




# =========================
# Inputs / Outputs
# =========================
# ここはあなたの環境に合わせて修正
XLSX_PATH = r"C:\Users\user\Desktop\PREFLOP_GAME_FOR_BEGINNERS-INTERMEDIATEコピー.xlsx"
SHEET_NAME = "zeros_range"  # ←実際のシート名に変える

OUT_DIR = Path(r"C:\MyPokerApp\tools\out")
OUT_JSON = OUT_DIR / "ref_colors_rgb.json"

# =========================
# Legend cell mapping
# =========================
# 本来は config.py にあるはず。あるなら import で持ってくるのが正しい。
# 例: from config import REF_COLOR_CELLS
try:
    from config import REF_COLOR_CELLS  # type: ignore
except Exception:
    # config から取れないなら、ここに暫定で最小例を置く（あなたの実物に置換すること）
    REF_COLOR_CELLS: Dict[str, Dict[str, str]] = {
        # "OR": {"RAISE_3BB": "D144", ...},
    }


# =========================
# Helpers
# =========================
_RGB6 = re.compile(r"^[0-9A-Fa-f]{6}$")


def _is_rgb_literal(s: str) -> bool:
    s = (s or "").strip()
    return bool(_RGB6.match(s))


def _read_fill_rgb(ws: Worksheet, a1: str) -> Optional[str]:
    """
    セルの fill.fgColor から RGB(6桁) を返す。
    取得できない/無効なら None。
    """
    cell = ws[a1]
    fill = cell.fill

    if fill is None or fill.patternType in (None, "none"):
        return None

    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None

    # openpyxl は rgb が ARGB("FF112233") で来ることが多い
    if getattr(fg, "type", None) == "rgb" and getattr(fg, "rgb", None):
        argb = str(fg.rgb).upper()
        rgb = argb[-6:]  # ARGB -> RGB
    else:
        # theme / indexed / auto はここでは解決しない（凡例生成ならスキップでOK）
        return None

    # “偽黒” を無効扱い（必要ならここは条件を調整）
    if rgb == "000000":
        return None

    return rgb


# =========================
# Main
# =========================
def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    ws_legend = wb["zeros_range"]   # REF_COLOR_CELLS の色凡例はここ
    ws_data   = wb["Datasheet"]     # AA_SEARCH_RANGES の探索はここ（このツールで使うなら）

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. sheets={wb.sheetnames}")

    ws = wb["zeros_range"]

    ref_colors_rgb: Dict[str, Dict[str, str]] = {}
    rows: List[Dict[str, str]] = []

    # 衝突検出用：rgb -> [(kind, raw_label), ...]
    rgb_to_labels: Dict[str, List[Tuple[str, str]]] = {}

    for kind, mapping in REF_COLOR_CELLS.items():
        kind_u = (kind or "").strip()
        out: Dict[str, str] = {}

        for raw_label, v in mapping.items():
            raw = (raw_label or "").strip()
            val = str(v).strip()

            if _is_rgb_literal(val):
                rgb = val.upper()
                source = "rgb_literal"
            else:
                # 値が空のセルは凡例として扱わない（装飾だけ残ってるノイズ回避）
                cell = ws[val]
                if cell.value is None or str(cell.value).strip() == "":
                    continue

                rgb = _read_fill_rgb(ws_legend, val)
                if rgb is None:
                    continue
                source = val  # A1番地を記録

            out[raw] = rgb
            rows.append({"kind": kind_u, "raw_label": raw, "rgb": rgb, "source": source})
            rgb_to_labels.setdefault(rgb, []).append((kind_u, raw))

        ref_colors_rgb[kind_u] = out

    print(
        f"[make_color_map] kinds={len(ref_colors_rgb)} rows={len(rows)} "
        f"unique_rgb={len(rgb_to_labels)}"
    )

    OUT_JSON.write_text(
        json.dumps(ref_colors_rgb, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[make_color_map] wrote: {OUT_JSON}")


if __name__ == "__main__":
    main()
