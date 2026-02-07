# excel_range_repository.py（末尾でもOK。既存クラスの外に dataclass を置いて、クラスにメソッド追加）
from __future__ import annotations

import logging
logger = logging.getLogger(__name__)

import re
from dataclasses import dataclass
from typing import Any, List, Tuple
from openpyxl.workbook.workbook import Workbook

RANKS = ["A","K","Q","J","T","9","8","7","6","5","4","3","2"]

def _rank_index(r: str) -> int:
    return RANKS.index(r)

@dataclass(frozen=True)
class RangeCellView:
    label: str      # Excelセルの表示（例: "AKs"）
    bg_rgb: str     # "RRGGBB"（"#"なし）

@dataclass(frozen=True)
class RangeGridView:
    kind: str
    pos: str
    sheet_name: str
    cells: List[List[RangeCellView]]  # 13x13
    aa_addr: str
    top_left: Tuple[int, int]         # (row, col)


def _normalize_hand_to_key(hand: str) -> str:
    """
    hand が "AKs"/"AKo"/"AA" などの既存キーの場合はそのまま。
    hand が "KsJc" / "AsKd" など2枚表記(4文字)の場合は "KJO" / "AKO" に正規化。

    NOTE:
    - 返り値は内部処理用に大文字化する（"S"/"O"）。
    """
    h = hand.strip()

    # 既に "AKs" / "KQo" / "AA" 形式っぽい
    if len(h) in (2, 3):
        return h.upper()

    # "KsJc" など（4文字想定: RankSuit + RankSuit）
    if len(h) == 4:
        r1, s1, r2, s2 = h[0].upper(), h[1].lower(), h[2].upper(), h[3].lower()

        # pair
        if r1 == r2:
            return f"{r1}{r2}"

        i1, i2 = _rank_index(r1), _rank_index(r2)
        hi, lo = (r1, r2) if i1 < i2 else (r2, r1)
        suited = (s1 == s2)
        return f"{hi}{lo}{'S' if suited else 'O'}"

    raise ValueError(f"Unrecognized hand format: {hand!r}")


def _expected_cell_label_from_hand_key(hand_key: str) -> str:
    """
    新Excelのセル内表示は末尾の s/o が無い想定。
    - "AKS"/"AKO" -> "AK"
    - "AA" -> "AA"
    """
    hk = hand_key.strip().upper()
    if len(hk) == 2:
        return hk
    if len(hk) == 3 and hk[2] in ("S", "O"):
        return hk[:2]
    raise ValueError(f"Unrecognized hand_key for label: {hand_key!r}")


def _hand_key_to_rc(hand_key: str) -> Tuple[int, int]:
    """
    hand_key: "AKS" / "AKO" / "AA"
    returns: (r0,c0) in [0..12]
      - diagonal: pair
      - upper triangle: suited
      - lower triangle: offsuit

    グリッド仕様:
      上三角 = suited, 下三角 = offsuit, 対角 = pair
    """
    hk = hand_key.strip().upper()

    # pair
    if len(hk) == 2:
        r = hk[0]
        i = _rank_index(r)
        return (i, i)

    # non-pair
    if len(hk) == 3 and hk[2] in ("S", "O"):
        r1, r2, so = hk[0], hk[1], hk[2]
        i1, i2 = _rank_index(r1), _rank_index(r2)

        # 強い方(小さい index)を hi
        hi_r, lo_r = (r1, r2) if i1 < i2 else (r2, r1)
        hi_i, lo_i = _rank_index(hi_r), _rank_index(lo_r)

        if so == "S":
            # 上三角
            return (hi_i, lo_i)
        else:
            # 下三角（suited 座標の転置）
            return (lo_i, hi_i)

    raise ValueError(f"Unrecognized hand_key: {hand_key!r}")


# =========================
# Position normalization (anchor search)
# =========================

_POS_NORM_RE = re.compile(r"[^A-Z0-9]+")


def _norm_pos_text(x: Any) -> str:
    """
    posセル探索用の正規化。
    - 大文字化
    - 英数字以外（空白/改行/記号/_ 等）を除去
    例:
      "BB vs SB" -> "BBVSSB"
      "BBvsSB "  -> "BBVSSB"
    """
    if x is None:
        return ""
    return _POS_NORM_RE.sub("", str(x).strip().upper())


# --- ref color parsing helpers (module-level) ---
_HEX6_RE = re.compile(r"^[0-9a-fA-F]{6}$")
_HEX8_RE = re.compile(r"^[0-9a-fA-F]{8}$")
_CELL_RE = re.compile(r"^[A-Za-z]{1,3}[0-9]{1,7}$")  # A1形式ざっくり

def _normalize_rgb(s: str) -> str | None:
    t = (s or "").strip()
    if not t:
        return None
    if t.startswith("#"):
        t = t[1:]
    if _HEX8_RE.match(t):
        return t[-6:].upper()  # ARGB -> RGB
    if _HEX6_RE.match(t):
        return t.upper()
    return None


def _is_cell_addr(s: str) -> bool:
    return bool(_CELL_RE.match((s or "").strip()))



# =========================
# Anchor match model
# =========================

@dataclass(frozen=True)
class AnchorMatch:
    pos_cell_addr: str
    pos_row: int
    pos_col: int
    aa_row: int
    aa_col: int
    aa_addr: str


# =========================
# Repository
# =========================

class ExcelRangeRepository:
    """
    Excelレンジ表を参照する Repository（posセル起点）。

    仕様（以前の設計を優先）:
    1) AA_SEARCH_RANGES[kind] 内で pos名(EP/MP/CO/BTN...) を検索
    2) posセルから (down=+3, left=-2) のセルが "AA"
    3) AAセルから GRID_TOPLEFT_OFFSET で 13x13 グリッド左上を求める
    4) hand_key -> (r0,c0) に変換して、13x13 内の該当セルを直接参照
       - そのセル色を読み、見本色と照合しタグを返す
       - 無色/不一致は "FOLD"

    見本色:
    - kind ごとに config で固定セル番地を持つ（ref_color_cells[kind][tag] = "H25" など）
    """

    def __init__(
        self,
        wb: Workbook,
        sheet_name: str,
        aa_search_ranges: Dict[str, str],             # kind -> A1 range
        grid_topleft_offset: Tuple[int, int],         # AA -> grid top-left (dr, dc)
        ref_color_cells: Dict[str, Dict[str, str]],   # kind -> tag -> "H25"
        enable_debug: bool = False,
    ) -> None:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}. Available={wb.sheetnames}")

        self.wb: Workbook = wb
        self.ws: Worksheet = wb[sheet_name]

        self.aa_search_ranges = dict(aa_search_ranges)
        self.grid_topleft_offset = tuple(grid_topleft_offset)
        self.ref_color_cells = dict(ref_color_cells)
        self.enable_debug = enable_debug

        # (kind, pos) -> AnchorMatch
        self._anchor_cache: Dict[Tuple[str, str], AnchorMatch] = {}

        # kind -> tag -> rgb
        self._ref_color_cache: Dict[str, Dict[str, str]] = {}
        
        self.debug_anchor_cache_hits = False

    # =========================
    # small safe getter (for debug only)
    # =========================
    def _safe_getattr(self, obj, name: str):
        try:
            return getattr(obj, name)
        except Exception as e:
            return f"<err:{e}>"

    # =========================
    # Anchor (pos -> AA)
    # =========================

    def find_anchor_by_pos(self, kind: str, pos: str) -> AnchorMatch:
        cache_key = (kind, pos)

        if cache_key in self._anchor_cache:
            m = self._anchor_cache[cache_key]
            # ★cache-hitはログ出さない（必要なら下のフラグで出せる）
            if self.enable_debug and getattr(self, "debug_anchor_cache_hits", False):
                print(
                    f"[REPO][ANCHOR] cached pos_cell={m.pos_cell_addr} -> AA={m.aa_addr} "
                    f"(kind={kind} pos={pos})",
                    flush=True,
                )
            return m


        if kind not in self.aa_search_ranges:
            raise KeyError(
                f"AA search range not defined for kind={kind}. "
                f"Defined kinds={list(self.aa_search_ranges.keys())}"
            )

        a1_range = self.aa_search_ranges[kind]
        candidates: list[AnchorMatch] = []

        pos_norm = _norm_pos_text(pos)

        for row in self.ws[a1_range]:
            for cell in row:
                if _norm_pos_text(cell.value) != pos_norm:
                    continue

                pr, pc = cell.row, cell.column
                aa_r, aa_c = pr + 3, pc - 2
                aa_cell = self.ws.cell(row=aa_r, column=aa_c)

                aa_val = "" if aa_cell.value is None else str(aa_cell.value).strip().upper()
                if aa_val != "AA":
                    if self.enable_debug:
                        print(
                            f"[REPO][ANCHOR] pos found at {cell.coordinate} but AA check failed "
                            f"(expected AA at {aa_cell.coordinate}, got {aa_cell.value!r})",
                            flush=True,
                        )
                    continue

                candidates.append(
                    AnchorMatch(
                        pos_cell_addr=cell.coordinate,
                        pos_row=pr,
                        pos_col=pc,
                        aa_row=aa_r,
                        aa_col=aa_c,
                        aa_addr=aa_cell.coordinate,
                    )
                )

        if not candidates:
            raise ValueError(
                f"Anchor not found for kind={kind}, pos={pos} within range={a1_range}. "
                f"(pos cell '{pos}' not found OR AA offset cell not 'AA')"
            )

        candidates.sort(key=lambda m: (m.pos_row, m.pos_col))
        chosen = candidates[0]
        self._anchor_cache[cache_key] = chosen

        if self.enable_debug:
            print(
                f"[REPO][ANCHOR] chosen pos_cell={chosen.pos_cell_addr} -> AA={chosen.aa_addr} "
                f"(kind={kind} pos={pos})",
                flush=True,
            )

        return chosen

    def list_positions(self, kind: str) -> list[str]:
        """
        AA_SEARCH_RANGES[kind] 内を走査して、
        「posセル + (down=+3,left=-2) が AA」になっている pos を列挙する。

        目的：generator側で pos をハードコードせず、Excelに存在するposだけ使う。
        """
        if kind not in self.aa_search_ranges:
            raise KeyError(
                f"AA search range not defined for kind={kind}. "
                f"Defined kinds={list(self.aa_search_ranges.keys())}"
            )

        a1_range = self.aa_search_ranges[kind]
        found: list[tuple[int, int, str]] = []

        for row in self.ws[a1_range]:
            for cell in row:
                val = cell.value
                if val is None:
                    continue

                pos_text = str(val).strip()
                if not pos_text:
                    continue

                pr, pc = cell.row, cell.column
                aa_r, aa_c = pr + 3, pc - 2
                if aa_c <= 0:
                    continue

                aa_cell = self.ws.cell(row=aa_r, column=aa_c)
                aa_val = "" if aa_cell.value is None else str(aa_cell.value).strip().upper()
                if aa_val != "AA":
                    continue

                found.append((pr, pc, pos_text))

        found.sort(key=lambda x: (x[0], x[1]))

        # 重複除去（同じ表示のposが複数箇所にあるケースに備える）
        uniq: list[str] = []
        seen: set[str] = set()
        for _, _, pos_text in found:
            key = _norm_pos_text(pos_text)  # 既存の正規化を利用
            if not key or key in seen:
                continue
            seen.add(key)
            uniq.append(pos_text)

        return uniq
    

    # =========================
    # Grid addressing
    # =========================

    def get_grid_top_left(self, kind: str, pos: str) -> Tuple[int, int]:
        """
        AAアンカーからグリッド左上(top-left)の座標(row,col)を返す。
        """
        anchor = self.find_anchor_by_pos(kind, pos)
        dr, dc = self.grid_topleft_offset
        return anchor.aa_row + dr, anchor.aa_col + dc

    def get_cell_value_at_grid(self, kind: str, pos: str, r0: int, c0: int) -> Any:
        top_r, top_c = self.get_grid_top_left(kind, pos)
        return self.ws.cell(row=top_r + r0, column=top_c + c0).value

    def get_cell_fill_rgb_at_grid(self, kind: str, pos: str, r0: int, c0: int) -> str:
        top_r, top_c = self.get_grid_top_left(kind, pos)
        cell = self.ws.cell(row=top_r + r0, column=top_c + c0)
        return self._read_fill_rgb(cell)

    # =========================
    # Reference colors (fixed cells by kind)
    # =========================

    def get_ref_colors(self, kind: str) -> Dict[str, str]:
        """
        kind ごとの見本色(tag -> RGB)を返す（キャッシュあり）。

        REF_COLOR_CELLS は「RGB直書き」or「セル番地」の両方を許可する：
          - RGB:  "f4cccc" / "#f4cccc" / "FFf4cccc"
          - A1 :  "D144" のようなセル番地（黒やテーマ色など例外用）
        """
        kind_u = (kind or "").strip().upper()

        if kind_u in self._ref_color_cache:
            return self._ref_color_cache[kind_u]

        if kind_u not in self.ref_color_cells:
            raise KeyError(
                f"REF_COLOR_CELLS not defined for kind={kind_u}. "
                f"Defined kinds={list(self.ref_color_cells.keys())}"
            )

        result: Dict[str, str] = {}
        mapping = self.ref_color_cells[kind_u]

        for tag, raw in mapping.items():
            raw_s = str(raw).strip()

            # 1) RGB直指定
            rgb = _normalize_rgb(raw_s)
            if rgb is not None:
                result[tag] = rgb
                continue

            # 2) セル番地
            if _is_cell_addr(raw_s):
                cell = self.ws[raw_s]
                rgb_read = self._read_fill_rgb(cell)  # 既存の色読み
                rgb2 = _normalize_rgb(rgb_read)
                if rgb2 is None:
                    raise ValueError(
                        f"Could not read RGB from cell {raw_s} for kind={kind_u} tag={tag}. "
                        f"Read='{rgb_read}'. Consider specifying RGB directly in REF_COLOR_CELLS."
                    )
                result[tag] = rgb2
                continue

            # 3) 不正値
            raise ValueError(
                f"Invalid REF_COLOR_CELLS value: kind={kind_u} tag={tag} value={raw_s!r} "
                f"(expected RGB hex like 'f4cccc' or cell addr like 'D144')"
            )

        self._ref_color_cache[kind_u] = result
        return result


    # =========================
    # Color reader
    # =========================

    def _read_fill_rgb(self, cell) -> str:
        """
        openpyxl Cell の塗りつぶし色(RGB)を "RRGGBB" で返す。
        塗りつぶし無し/取得不能は ""。

        重要:
        - patternType が無い/none の場合は "" にする（無色の誤一致を防ぐ）
        - theme/indexed はまず "" 扱い（必要なら後で拡張）
        """
        fill = getattr(cell, "fill", None)
        if fill is None:
            return ""

        pattern = getattr(fill, "patternType", None)
        if pattern is None or str(pattern).lower() in ("none", "null"):
            return ""

        fg = getattr(fill, "fgColor", None) 
        if fg is None:
            return ""

        fg_type = getattr(fg, "type", None)
        if fg_type not in (None, "rgb"):
            return ""

        rgb = getattr(fg, "rgb", None)
        if not rgb:
            return ""

        rgb = str(rgb).upper()
        if len(rgb) == 8:
            rgb = rgb[-6:]
        elif len(rgb) != 6:
            return ""

        return rgb

    # =========================
    # Main API: tag lookup
    # =========================

    def get_tag_for_hand(self, kind: str, position: str, hand: str) -> Tuple[str, Dict[str, Any]]:
        """
        hand_key -> (r0,c0) -> グリッド直接参照 -> fill色でタグ判定。
        追加仕様：
        - 「セル値（ハンド名）と色」が両方揃ったときだけ有効
          文字列のみ / 色のみ は “色なし” と同じ扱い（= FOLD）
        """
        debug: Dict[str, Any] = {"kind": kind, "position": position, "hand_in": hand}
        hand_key = _normalize_hand_to_key(hand)
        r0, c0 = _hand_key_to_rc(hand_key)
        expected_label = _expected_cell_label_from_hand_key(hand_key)
        debug.update({"hand_key": hand_key, "r0": r0, "c0": c0, "expected_label": expected_label})

        # 1) グリッド左上
        top_r, top_c = self.get_grid_top_left(kind, position)
        debug["grid_topleft"] = (top_r, top_c)

        target_row = top_r + r0
        target_col = top_c + c0
        cell = self.ws.cell(row=target_row, column=target_col)

        debug["target_cell_rc"] = (target_row, target_col)
        debug["target_cell_a1"] = cell.coordinate
        debug["cell_value"] = cell.value

        # ★セル値チェック（色だけの凡例セルなどを除外）
        cell_text = "" if cell.value is None else str(cell.value).strip().upper()
        debug["cell_text_norm"] = cell_text

        if cell_text != expected_label:
            # 文字列のみ・色のみは「色なし」と同じ扱いにする
            debug["cell_rgb"] = ""  # 強制的に無色扱い
            ref = self.get_ref_colors(kind)
            debug["ref_colors"] = ref
            debug["tag"] = "FOLD"
            debug["rejected_reason"] = "cell_label_mismatch_or_blank"
            return "FOLD", debug

        # 2) 対象セルの色（ここまで来たら “文字＋色” の色を見る）
        rgb = self._read_fill_rgb(cell)
        debug["cell_rgb"] = rgb

        # 3) 見本色と照合
        ref = self.get_ref_colors(kind)  # tag -> rgb
        debug["ref_colors"] = ref

        if not rgb:
            debug["tag"] = "FOLD"
            debug["rejected_reason"] = "no_fill_color"
            return "FOLD", debug

        for tag, ref_rgb in ref.items():
            if rgb == ref_rgb and ref_rgb:
                debug["tag"] = tag
                return tag, debug

        debug["tag"] = "FOLD"
        debug["unmatched_rgb"] = rgb
        return "FOLD", debug

    def hand_to_grid_rc(self, card1: str, card2: str) -> Tuple[int, int]:
        """
        ("Ks","Jc") -> (r,c) in 0..12 のグリッド座標。
        - ペア: 対角
        - suited: 対角より上（row < col）
        - offsuit: 対角より下（row > col）
        """
        r1, s1 = card1[0], card1[1]
        r2, s2 = card2[0], card2[1]
        i1 = _rank_index(r1)
        i2 = _rank_index(r2)

        if r1 == r2:
            return (i1, i1)

        suited = (s1 == s2)
        hi = min(i1, i2)  # indexが小さいほど高ランク
        lo = max(i1, i2)

        return (hi, lo) if suited else (lo, hi)

    def get_range_grid_view(self, kind: str, pos: str, size: int = 13):
        """
        表示専用：該当レンジ表の 13x13 を (label, bg_rgb) で返す。
        アンカー探索は1回だけにして、ログ連発と無駄呼び出しを防ぐ。
        """
        anchor = self.find_anchor_by_pos(kind, pos)

        # ★ここが重要：get_grid_top_left() を呼ばずに top-left を計算（find_anchorの再実行を防ぐ）
        dr, dc = self.grid_topleft_offset
        top_r = anchor.aa_row + dr
        top_c = anchor.aa_col + dc

        cells = []
        for r0 in range(size):
            row_cells = []
            for c0 in range(size):
                cell = self.ws.cell(row=top_r + r0, column=top_c + c0)

                v = cell.value
                label = "" if v is None else str(v).strip()

                rgb = self._read_fill_rgb(cell)  # あなたの既存関数を直接使う
                rgb = (rgb or "FFFFFF")[-6:].upper()

                row_cells.append(RangeCellView(label=label, bg_rgb=rgb))
            cells.append(row_cells)

        return RangeGridView(
            kind=kind,
            pos=pos,
            sheet_name=self.ws.title,
            cells=cells,
            aa_addr=anchor.aa_addr,
            top_left=(top_r, top_c),
        )


